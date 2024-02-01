import os
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook

from romanian_months import *


#To avoid repetition/clutter of bills
if os.path.exists("finaldata.csv"):
    os.remove("finaldata.csv")


UserName = input("Nume angajat(Popescu Ion): ")
ExpenseCode = input("Cod decontare (sau explicatie cost):")

# Initialize variables
InvoiceNumbers = []
Issuers = []
Prices = []
Dates = []

# Loop through PDF files
for filename in os.listdir("."):
    if filename.endswith(".pdf"):
        pdf_path = os.path.join(".", filename)
        reader = PdfReader(pdf_path)

        # Extract text from all pages
        text = ""
        for page in reader.pages:
            text += page.extract_text()


        # Initialize our needed values
        Price = None
        InvoiceNo = None
        Issuer = None
        Date = None

        printed_lines = set()

        for line in text.split("\n"):
            stripped_line = line.strip()

            if "(RON)" in stripped_line and "including" in stripped_line and stripped_line not in printed_lines:
                Price = stripped_line.replace("T otal  including  V A T  (RON): ", "")  
                Prices.append(Price)
                printed_lines.add(Price)

            if "Invoice  no." in stripped_line and stripped_line not in printed_lines:
                InvoiceNo = stripped_line.replace("Invoice  no.  ", "")
                InvoiceNumbers.append(InvoiceNo)
                printed_lines.add(InvoiceNo)

            if "Issued  on  behalf  of  " in stripped_line and stripped_line not in printed_lines:
                Issuer = stripped_line.replace("Issued  on  behalf  of  ", "")
                Issuer = Issuer.split("by  Bolt  Operations", 1)
                Issuer = Issuer[0]
                Issuers.append(Issuer)
                printed_lines.add(Issuer)

            if "Date:  " in stripped_line and stripped_line not in printed_lines:
                Date = stripped_line.replace("Date:  ", "")
                Dates.append(Date)
                printed_lines.add(Date)
        

        # Check if all information is available before writing to CSV
        if all([InvoiceNo, Price, Issuer, Date]):
            csv_data = [
                ["InvoiceNo", "Price", "Issuer", "Date"],
                [InvoiceNo, Price, Issuer, Date]
            ]
            
            # Append data to the CSV file
            with open("finaldata.csv", "w", newline="\n") as finaldata:
                csv_writer = csv.writer(finaldata)
                csv_writer.writerows(csv_data)


# From this part on, the data is written in Excel
                
for filename in os.listdir("."):
    if filename.endswith(".xlsx"):
        xlspath = os.path.join(".",filename)

wb = load_workbook(xlspath)
ws = wb.active


# Single values
ws["D6"] = formatted_date
ws["D4"] = UserName

# Usually multiple values, unless you have just one bill
start_invoice = "B13"
start_issuer = "D13"
start_price = "G13"
start_date = "C13"
start_info = "H13"


for i, invoice in enumerate(InvoiceNumbers):

    ws[start_invoice] = invoice
    ws[start_info] = ExpenseCode
    start_invoice = f'B{13 + i + 1}'
    start_info = f'H{13 + i + 1}'
    

for i, issuer in enumerate(Issuers):
    
    ws[start_issuer] = issuer
    start_issuer = f'D{13 + i + 1}'

for i, price in enumerate(Prices):
    
    ws[start_price] = float(price)
    start_price = f'G{13 + i + 1}'

for i, date in enumerate(Dates):
    
    ws[start_date] = date
    start_date = f'C{13 + i + 1}'


# The proper file name is written automatically, one less thing to manually worry about :) 
wb.save("Decont " + current_month_ro + " " + current_year + " " + UserName + ".xlsx")
